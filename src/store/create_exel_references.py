'''ОБЩАЯ ИНФОРМАЦИЯ'''
data = {
    # Признак ИС/ИР
    'sign': [
        'ИС', 
        'ИР',
        'ИС/ИР'
    ],
    # Этап эксплуатации
    'stage': [
        'Разработка',
        'Тестовая эксплуатация (ТЭ)',
        'Опытная эксплуатация (ОЭ)',
        'Опытно-промышленная эксплуатация (ОПЭ)',
        'Промышленная эксплуатация (ПЭ)'
    ],
    # Вид эксплуатации
    'view_operation': [
        'Инфраструктурная',
        'Производственная',
        'Непроизводственная',
        'Информационная безопасность',
        'Экономическая безопасность'
    ],


    # '''СЕРВИСНОЕ ОБСЛУЖИВАНИЕ'''
    # Доступность
    'availability': [
        'Без доступа',
        '7х24'
    ],
    # Сетевая конфигурация
    'network_configuration': [
        'Клиент-сервер (в пределах одной территории)',
        'Распределенная серверная группировка',
        'Распределенная, с одной серверной группировкой',
        'Распределенная, с двумя серверными группировками'
    ],
    # Аппаратное обеспечение
    'hardware': [
        'Виртуальный сервер',
        'Виртуальный сервер vxrailserver',
        'Виртуальный сервер master',
        'Виртуальный сервер VMware ESXi',
        'Физический сервер',
        'Физический сервер СУБД',
        'Физический сервер SQL',
        'Сервер DELL PowerEdge R640',
        'Сервер кластера виртуализации',
        'Кластер серверов БД',
        'Кластер серверов БД Oracle',
        'Кластер сервера Sun Enterprise T5440',
        'Кластер серверов управления (сервер мониторинга)',
        'Узел кластеров',
        'Сервер РК',
        'Агент РК',
        'VMware Virtual Machine',
        'Oracle',
        'Мини-ПК для телевизоров',
        'Шасси блейд-серверов VxRail G410',
        'Шасси блейд-серверов VxRail G560',
        'Блейд-сервер VxRail',
        'Система хранения данных',
        'Дисковые библиотеки',
        'HP ProLiant DL360 Gen',
        'Сервер управления Системой (сервер мониторинга)',
        'Центральный сервер Системы (CUCM Publisher), виртуальный',
        'Голосовой шлюз',
        'Объекты удаленного управления серверами производства Hewlett Packard (iLO), Dell (iDRAC), SUN/Oracle (ILOM), IBM/Lenovo (RSA)',
        'Объекты удаленного управления операционными системами Microsoft Windows Server',
        'Объекты удаленного управления операционными системами UNIX/Linux',
        'Активное сетевое оборудование ЛВС/ГВ'
    ],
    # Программное обеспечение
    'software': [
        '.NET',
        '1С: PM Управление проектами',
        '1С: Предприятие 8.3. КОРП',
        'Apache',
        'Apache Tomcat',
        'Apache Tomcat 7.0.39',
        'Apache Tomcat 9.0.14',
        'ArcGIS for Desktop 10.3',
        'ArcGIS for Server 10.3',
        'ASP.NET Core 6.0 ',
        'Astra Linux Common Edition 2.12',
        'Astra Linux Special Edition 1.7',
        'CentOS 7.9',
        'Cisco IOS',
        'Cisco IP Communicator',
        'Cisco Unified Communications Manager Version 11.5',
        'DIRECTUM 5.7',
        'DIRECTUM 5.8.2',
        'Docker Compose 1.24.1',
        'Docker Engine 19.03.2',
        'Docker Engine 20.10.21',
        'Eclipse',
        'ELMA Community Edition',
        'EMC Isilon OneFS 8.0.0.5',
        'EMC Networker for UNIX Network Edition',
        'ESRI ArcSDE 10.0',
        'Exchange Server 2010 Enterprise',
        'FlexNet',
        'FlexNet Publisher 2021 R4 (11.18.3.0)',
        'GeoFrame',
        'Geoserver',
        'Geoserver 2.20.4',
        'Geovation',
        'GitLab Community Edition 15.3',
        'GitLab Runner 15.3',
        'Guardant',
        'Internet Information Server (IIS) 8.5',
        'Internet Information Server (IIS) 10.0',
        'Internet Information Server (IIS) 7.5.7600.16385',
        'Java 2 Enterprise Edition',
        'Java Standard Edition 1.6',
        'JBoss Application Server 6',
        'Kaspersky Security Center',
        'Kaspersky Security Center 10 ',
        'Keycloak 18.0.0',
        'MaxPatrol 8',
        'Microsoft .Net Framework 4.7.2',
        'Microsoft Office Professional Plus 2013',
        'Microsoft SharePoint 2016',
        'Microsoft SQL Server 2008 R2 Enterprise',
        'Microsoft SQL Server 2014 Standard',
        'Microsoft SQL Server 2016 Standard',
        'Microsoft SQL Server 2017 Express',
        'Microsoft SQL Server 2017 Standard',
        'Microsoft Visual Basic for Applications 7.1',
        'Microsoft Windows Server 2008 R2 Enterprise',
        'Microsoft Windows Server 2008 R2 Standard',
        'Microsoft Windows Server 2012 R2 Standard',
        'Microsoft Windows Server 2016',
        'Microsoft Windows Server 2016 Datacenter',
        'Microsoft Windows Server 2016 Standard',
        'Minio 1.13',
        'MS SCCM Remote Control',
        'MySQL 6.0',
        'NetWorker 9.2.0.3.Build.85',
        'NGINX 1.18',
        'Node.js v11.6.0',
        'Node.js v16.16.0',
        'Npm 6.5.0',
        'Npm 8.11.0',
        'OIS',
        'Open iT',
        'Oracle 11.2.0',
        'Oracle 11.2.0.1.0',
        'Oracle 11.2.0.4',
        'Oracle 11.2.0.4.0',
        'Oracle Database',
        'Oracle Database 11g Enterprise Edition',
        'Oracle Database 11g R2',
        'Paradigm',
        'Petrel',
        'Petrel-Studio',
        'PostGIS',
        'Postgres Pro',
        'Postgres Pro Enterprise',
        'PostgreSQL',
        'PostgreSQL 9.6',
        'PostgreSQL 11.12',
        'PostgreSQL 12',
        'PostgreSQL 12.0',
        'PostgreSQL 12.1',
        'PostgreSQL 14',
        'PostgreSQL 14.1',
        'PostgreSQL 14.2',
        'PostgreSQL Database Server 10.4',
        'Prime',
        'PRTG Network Monitor 15.1.14.1710',
        'Python 3.6',
        'Python 3.10',
        'QGIS 3.8',
        'Red Hat Enterprise Linux 5.7',
        'Red Hat Enterprise Linux 5.11',
        'Red Hat Enterprise Linux 6.5',
        'Red Hat Enterprise Linux 6.8',
        'Red Hat Enterprise Linux 6.9',
        'Red Hat Enterprise Linux 7.4',
        'Redmine 5.0',
        'SCCM 2007 R2',
        'Sentinel',
        'Solaris 10',
        'Sonatype Nexus Repository 3',
        'SUSE Linux Enterprise Server 12 SP1',
        'tNavigator',
        'Ubuntu Server 18.04 LTS',
        'VISAGE',
        'VMware ESXi 6.5',
        'VMware Horizon 7',
        'VMWare vSphere 6.5',
        'ИСУ ТТПК',
        'МАГМА',
        'ОС Solaris 10',
        'РН-КИМ',
        'РН-КИН',
        'САПСАН',
        'Система бюджетного управления (Разработчик ООО «СибирьСофтПроект»)'
    ],
    # Тип СУБД
    'type_subd': [
        'Не используется',
        'PostgreSQL',
        'Microsoft SQL Server',
        'MySQL',
        'Oracle',
    ],
    # Степень критичности
    'degree_criticality': [
        'Отсутствует',
        'Низкая',
        'Средняя',
        'Высокая'
    ],


    # '''УРОВЕНЬ ДОСТУПА ИНФОРМАЦИИ'''
    # Уровень конфиденциальности
    'level_privacy': [
        'Не категорируется',
        'Открытый',
        'Для внутреннего пользования',
        'Конфиденциальный'
    ],
    # 'file_status': [
    #     'Отсутствует',
    #     'В разработке',
    #     'На согласовании',
    #     'Согласован'
    # ],
    'platform_version': [
        'Нет',
        'Виртуальная платформа VxRail'
    ],
    'names_BM': [
        'MAP',
        'SDE',
        'tmn-tnnc-ksc.rosneft.ru',
        'tmn-tnnc-db.rosneft.ru',
        'tmn-tnnc-nas01.rosneft.ru',
        'tmn-tnnc-nas02.rosneft.ru',
        'tmn-tnnc-nas03.rosneft.ru',
        'tmn-tnnc-ts.rosneft.ru ',
        'tmn-tnnc-rn-lab',
        'TMN-DB-TNNC01MAP',
        'tnnc-rnlab-app01',
        'tmn-dba1.rosneft.ru',
        'tmn-webapp.rosneft.ru',
        'tnnc-bpm.rosneft.ru',
        'tnnc-bpm-test.rosneft.ru ',
        'tnnc-bpm-db.rosneft.ru',
        'tnnc-index.rosneft.ru',
        'T7Host1 (10.28.71.107)',
        'T7Host0 (10.28.71.106)',
        'TNNC-TS-UPR.rosneft.ru',
        'TNNC-DB-RNLAB',
        'TNNC-GIS-APP',
        'TNNC-UPP.rosneft.ru',
        'TNNC-UPP-test.rosneft.ru',
        'TNNC-OPER.rosneft.ru',
        'TNNC-OPER-test.rosneft.ru',
        'TNNC-MIC.rosneft.ru',
        'TNNC-MIC-test.rosneft.ru',
        'TNNC-PUED-DB.rosneft.ru   ',
        'TNNC-PUED-DB-test.rosneft.ru',
        'tnnc-upr.rosneft.ru',
        'TNNC-BALANCE',
        'TNNC-DEMIURG-APP01',
        'TNNC-MOB-BD',
        'TMN-TNNC-DEPCAB',
        'TNNC-BSRVC2.rosneft.ru',
        'TNNC-BSRVC1.rosneft.ru',
        'TNNC-SRVCR1.rosneft.ru',
        'TNNC-SRVCR2.rosneft.ru',
        'TNNC-SRVCDB.rosneft.ru',
        'tmn-tnnc-dev.rosneft.ru',
        'Tnnc-info.rosneft.ru ',
        'tmn-tnnc-oauth.rosneft.ru',
        'TNNC-SVP-AS01',
        'TNNC-SVP-AS02',
        'TNNC-SVP-AS03',
        'TNNC-SVP-AS04',
        'TNNC-SVP-AS06',
        'TNNC-TS-SVP',
        'TNNC-OUN-AS',
        'TNNC-OUN-APP01',
        'TNNC-OUN-APP02',
        'TNNC-OUN-APP03',
        'TNNC-OUN-APP04',
        'TNNC-OUN-TST01',
        'TNNC-OUN-TST02',
        'TNNC-SBI-AS01',
        'tnnc-ucto-as01'
    ],
    'unit_measure': [
        'Гб',
        'Тб'
    ],
    'doc_type': [
        # 'Лист изменения',
        'ВТР',
        'ФТТ/ТЗ',
        'Перечень информации, используемой в бизнес-процессе',
        'Решение о категорировании информации (РОК)',
        'Технический проект (ТПр)',
        'Программа и методика испытаний (ПиМИ)',
        'Программа и методика испытаний на соответствие требованиям ИБ',
        'Результаты проведенных испытаний (Протокол ПСИ)',
        'Протокол ОС/УС о готовности к старту этапа эксплуатации',
        'Протокол приема-передачи в сопровождение ИС/ИР',
        'Акт о готовности ИС/ИР к вводу в промышленную эксплуатацию',
        'Технический паспорт (ТПс)',
        'Регламент предоставления доступа (РПД)',
        'Реестр привилегированных учетных записей',
        'Руководство по обеспечению непрерывности (РОН)',
        'Инструкция пользователя',
        'Инструкция администратора',
        'Приказ',
    ],
    # 'doc_group': [
    #     'Проектная и управленческая документация',
    #     'Эксплуатационная документация'
    # ],
    # 'doc_status': [
    #     'Разработка',
    #     'Актуализация',
    #     'На согласовании',
    #     'Согласовано'
    # ],



}

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import TableList
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import get_column_letter
# from rich import print


def load_for_default_table():
    wb = Workbook()
    sheets_name = data.items()
    
    for name, vals in sheets_name:
        ws: Worksheet = wb.create_sheet(name)
        headers = ["id", "name"]
        items = [[idx, cel] for idx, cel in enumerate(vals)]
        rows = [headers] + items
        for row in rows:
            ws.append(row)
        letter = get_column_letter(2)
        cell_end = f"{letter}{len(rows)}"
        ws.column_dimensions[letter].width = 100
        # create a table
        tab = Table(displayName=f"Table_{name}", ref= F"A1:{cell_end}")
        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    wb.remove(wb["Sheet"])
    wb.save(r"src\store\default_references.xlsx")
    wb.close()
    print('"OK"')



if __name__ == "__main__":
    load_for_default_table()





        # obj = {}
        # obj["id"] = 
        # ws.append(vals)