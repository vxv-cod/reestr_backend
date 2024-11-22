import asyncio
import datetime
import inspect
import os
import sys
from loguru import logger
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.cell.cell import Cell
# from rich import print
from sqlalchemy import delete


if __name__ == "__main__":
    sys.path.insert(1, os.path.join(sys.path[0], '..'))


from apps.repositories.repo_SQL import SQLRepo
from apps.repositories.repo_uow import UnitOfWork
from store.create_exel_references import data as default_data
from utils.funcs import datetime_to_string, write_json, write_py



# def load_checklist(filename: str = 'src\store\checklist.xlsx', sheet: str = 'Лист1', fist_row: int = 16):

#     '''Загружаем из файла объект екселя'''
#     wb = load_workbook(filename = filename)
#     values = wb[sheet].values
#     rows = [row for row in values]
#     '''Пересобираем список без первых строк'''
#     rows = rows[fist_row - 1 : ]
#     '''Пересобираем список без первой колонки'''
#     rows = [list(row[1:]) for row in rows if row[1] != None]
#     sign = default_data["sign"]
#     result = []
#     for idx, row in enumerate(rows):
#         logger.debug(f"{row = }")
#         obj = {}
#         obj["id"] = idx
#         obj["name"] = row[0]
#         obj["sign_id"] = sign.index(row[1])
#         obj["test"] = 1 if row[2] == "Требуется" else 0
#         obj["experience"] = 1 if row[3] == "Требуется" else 0
#         obj["pilot_industrial"] = 1 if row[4] == "Требуется" else 0
#         obj["industrial"] = 1 if row[5] == "Требуется" else 0
#         result.append(obj)
#     # logger.debug(f"{sign = }")
#     logger.debug("OK")
#     return result


# def load_for_default_table(sheet: str):
#     filename: str = 'src\store\default_table.xlsx'
#     '''Загружаем из файла объект екселя'''
#     wb = load_workbook(filename = filename, data_only=True)
#     values = wb[sheet].values
#     rows = [row for row in values]
#     result = []
#     for row in rows[1:]:
#         tmp = {}
#         for idx, cel in enumerate(rows[0]):
#             tmp[cel] = row[idx]
#         result.append(tmp)
#     print(result)
#     return result




'''reestr_is'''
def get_exel(filename: str):
    wb = load_workbook(filename = filename, data_only=True)
    sheets = wb.sheetnames
    result = {}
    for name in sheets:
        all_rows = [row for row in wb[name].values]
        header = all_rows[0]
        rows = all_rows[1:]
        tmp_list = []
        for row in rows:
            tmp = {}
            for idx, col in enumerate(header):
                '''Выбираем все строки без колонки id, 
                чтобы она создалась сама для ативации sequence'''
                if col != 'id':
                    tmp[col] = row[idx]
            tmp_list.append(tmp)
        result[name] = tmp_list
    print(result)
    return result







async def sqlgo(table, data, uow: UnitOfWork):
    uow_attr: SQLRepo = getattr(uow, table)
    await uow.session.execute(delete(uow_attr.model))
    await uow.session.flush()
    await uow_attr.add_list(data)
    await uow.session.flush()  



async def sqlmultisave(obj: dict, uow: UnitOfWork):
    for key, val in obj.items():
        if isinstance(val, list):
            await sqlgo(key, val, uow)
            for item in val:
                if isinstance(item, dict):
                    await sqlmultisave(item, uow)


async def load_db(data_json, uow: UnitOfWork = UnitOfWork()):
    async with uow:
        await sqlmultisave(data_json, uow)
        await uow.commit()   
        print('"Exel >>> DataBase"')



if __name__ == "__main__":
    # from rich import print
    # checklist = load_checklist()
    # '''Сохраняем в файл в store'''
    # write_json(checklist, "src/store/checklist")
    
    # data = load_for_default_table('doc_type')
    
    default_references = get_exel(r'src\store\default_references.xlsx')
    asyncio.run(load_db(default_references))

    test_table = get_exel(r'src\store\test_table.xlsx')
    asyncio.run(load_db(test_table))

    
 

    # write_json(data, "src/store/reestr")
    
    
    
    